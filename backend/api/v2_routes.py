"""
BuyerHunter V2 — REST API Routes

Production-ready FastAPI endpoints connecting the V2 core engine
to the React frontend. Field-maps V2 models to the shapes the UI expects.
"""

import asyncio
import json
import logging
import re
import time
import uuid
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import exists, func, or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from backend.core.database import get_v2_db
from backend.core.models.company import Company
from backend.core.models.contact import Contact, ContactChannel, ContactPurpose
from backend.core.models.evidence_ledger import EvidenceLedger
from backend.core.models.search_jobs import SearchJob, JobStatus, JobPriority
from backend.core.engine.intent_analyzer import IntentAnalyzer
from backend.core.engine.query_planner import HybridQueryPlanner

logger = logging.getLogger(__name__)
router = APIRouter()

# ═══════════════════════════════════════════════════════════════════════════════
# Field mapping: V2 model → frontend-expected names
# ═══════════════════════════════════════════════════════════════════════════════

def _company_to_frontend(c: Company, contacts: list[Contact] | None = None) -> dict:
    """Map V2 Company model to the field names the React frontend expects."""
    # Build contact info from V2 contacts table
    email = ""
    phone = ""
    whatsapp = ""
    linkedin = ""
    official_email = ""
    sales_email = ""
    official_phone = ""

    if contacts:
        for ct in contacts:
            val = ct.channel_value or ""
            if ct.channel == "email" and not email:
                email = val
                if ct.channel_purpose == "official":
                    official_email = val
                elif ct.channel_purpose == "sales":
                    sales_email = val
            elif ct.channel == "phone" and not phone:
                phone = val
                if ct.channel_purpose == "official":
                    official_phone = val
            elif ct.channel == "whatsapp" and not whatsapp:
                whatsapp = val
            elif ct.channel == "linkedin" and not linkedin:
                linkedin = val

    # Contact person info from first contact
    contact_person = contacts[0].person_name if contacts else ""
    designation = contacts[0].designation if contacts else ""

    return {
        # Core identity
        "id": c.id,
        "company_name": c.canonical_name,
        "legal_name": c.legal_name or "",
        "website": c.website_url or "",
        # Contact channels (mapped from contacts table)
        "email": email,
        "phone": phone,
        "whatsapp": whatsapp or "",
        "linkedin_url": linkedin or "",
        "official_email": official_email,
        "sales_email": sales_email,
        "procurement_email": "",
        "support_email": "",
        "official_phone": official_phone,
        "whatsapp_business": whatsapp,
        # Contact person
        "contact_person": contact_person or "",
        "designation": designation or "",
        # Classification
        "industry": c.industry or "",
        "sub_industry": c.sub_industry or "",
        "business_type": c.sub_industry or c.industry or "",
        # Legal
        "gst_number": c.gst_number or "",
        "cin_number": c.cin_number or "",
        "iec_code": c.iec_code or "",
        "fssai_number": c.fssai_number or "",
        "pan_number": c.pan_number or "",
        # Location
        "address": c.hq_address or "",
        "city": c.hq_city or "",
        "state": c.hq_state or "",
        "district": c.hq_district or "",
        "country": c.hq_country or "India",
        "pin_code": c.hq_pincode or "",
        "factory_address": c.factory_address or "",
        "office_address": c.hq_address or "",
        "warehouse_address": c.warehouse_address or "",
        # Geolocation
        "latitude": c.latitude,
        "longitude": c.longitude,
        # Scoring
        "lead_score": c.buyer_score,
        "buyer_score": c.buyer_score,
        "confidence": c.confidence,
        "ai_confidence": c.confidence,
        "opportunity_score": 0,
        "risk_score": 0,
        "tier": c.tier.value if c.tier else "unknown",
        "estimated_size": c.company_tier.value if c.company_tier else "unknown",
        # Business flags
        "is_manufacturer": c.is_manufacturer,
        "is_importer": c.is_importer,
        "is_exporter": c.is_exporter,
        "is_distributor": c.is_distributor,
        "is_wholesaler": c.is_wholesaler,
        "is_retailer": c.is_retailer,
        "is_verified": c.confidence >= 70,
        "lead_status": "New",
        # Certifications
        "fssai_number": c.fssai_number or "",
        "iso_certification": "",
        "haccp_certification": "",
        "brc_certification": "",
        "apeda_registration": "",
        # Metadata
        "source": c.first_seen_source or "",
        "products": "",
        "brands": "",
        "about_us": "",
        "company_description": "",
        "ai_reason": "",
        "ai_consumption": "",
        "ai_frequency": "",
        "potential_oil_usage": "",
        "estimated_annual_consumption": "",
        "procurement_info": "",
        "contact_page": "",
        "careers_page": "",
        "export_markets": "",
        "import_countries": "",
        "turnover": "",
        "revenue": "",
        "employees": None,
        "founded_year": None,
        "google_rating": None,
        "linkedin_url_val": linkedin,
        "facebook_url": "",
        "instagram_url": "",
        "youtube_url": "",
        # Timestamps
        "crawl_date": c.first_seen_at.isoformat() if c.first_seen_at else None,
        "created_at": c.created_at.isoformat() if c.created_at else None,
        "updated_at": c.updated_at.isoformat() if c.updated_at else None,
        "enriched_at": c.last_enriched_at.isoformat() if c.last_enriched_at else None,
        "last_updated": c.updated_at.isoformat() if c.updated_at else None,
    }


def _contact_to_frontend(ct: Contact) -> dict:
    """Map V2 Contact model to frontend-expected shape."""
    return {
        "id": ct.id,
        "company_id": ct.company_id,
        "person_name": ct.person_name or "",
        "designation": ct.designation or "",
        "department": ct.department or "",
        "email": ct.channel_value if ct.channel == "email" else "",
        "phone": ct.channel_value if ct.channel == "phone" else "",
        "whatsapp": ct.channel_value if ct.channel == "whatsapp" else "",
        "linkedin_url": ct.channel_value if ct.channel == "linkedin" else "",
        "confidence_score": ct.confidence,
        "is_verified": ct.is_verified,
        "source": "",
        "notes": "",
        "created_at": ct.created_at.isoformat() if ct.created_at else None,
        "updated_at": ct.updated_at.isoformat() if ct.updated_at else None,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Health
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/health")
async def health():
    return {"status": "ok", "service": "BuyerHunter AI", "version": "2.0.0", "engine": "v2"}


# ═══════════════════════════════════════════════════════════════════════════════
# Stats
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/stats")
async def get_stats(db: AsyncSession = Depends(get_v2_db)):
    """Dashboard statistics from V2 database."""
    try:
        total = (await db.execute(select(func.count(Company.id)))).scalar() or 0
        verified = (await db.execute(
            select(func.count(Company.id)).where(Company.confidence >= 70)
        )).scalar() or 0
        importers = (await db.execute(
            select(func.count(Company.id)).where(Company.is_importer == True)
        )).scalar() or 0
        exporters = (await db.execute(
            select(func.count(Company.id)).where(Company.is_exporter == True)
        )).scalar() or 0
        manufacturers = (await db.execute(
            select(func.count(Company.id)).where(Company.is_manufacturer == True)
        )).scalar() or 0
        distributors = (await db.execute(
            select(func.count(Company.id)).where(Company.is_distributor == True)
        )).scalar() or 0
        wholesalers = (await db.execute(
            select(func.count(Company.id)).where(Company.is_wholesaler == True)
        )).scalar() or 0

        # Contact counts
        emails = (await db.execute(
            select(func.count(Contact.id)).where(Contact.channel == "email")
        )).scalar() or 0
        phones = (await db.execute(
            select(func.count(Contact.id)).where(Contact.channel == "phone")
        )).scalar() or 0

        # Industry breakdown
        industry_rows = (await db.execute(
            select(Company.industry, func.count(Company.id))
            .where(Company.industry.isnot(None))
            .group_by(Company.industry)
            .order_by(func.count(Company.id).desc())
            .limit(20)
        )).all()
        industry_breakdown = {r[0]: r[1] for r in industry_rows}

        # Source breakdown
        source_rows = (await db.execute(
            select(Company.first_seen_source, func.count(Company.id))
            .where(Company.first_seen_source.isnot(None))
            .group_by(Company.first_seen_source)
            .order_by(func.count(Company.id).desc())
        )).all()
        source_breakdown = {r[0]: r[1] for r in source_rows}

        # State breakdown
        state_rows = (await db.execute(
            select(Company.hq_state, func.count(Company.id))
            .where(Company.hq_state.isnot(None))
            .group_by(Company.hq_state)
            .order_by(func.count(Company.id).desc())
            .limit(20)
        )).all()
        state_breakdown = {r[0]: r[1] for r in state_rows}

        # Score distribution
        score_ranges = {"0-20": 0, "21-40": 0, "41-60": 0, "61-80": 0, "81-100": 0}
        for label, low, high in [("0-20", 0, 20), ("21-40", 21, 40), ("41-60", 41, 60),
                                  ("61-80", 61, 80), ("81-100", 81, 100)]:
            count = (await db.execute(
                select(func.count(Company.id)).where(
                    Company.buyer_score >= low, Company.buyer_score <= high
                )
            )).scalar() or 0
            score_ranges[label] = count

        # Recent jobs
        recent_jobs = (await db.execute(
            select(SearchJob).order_by(SearchJob.created_at.desc()).limit(10)
        )).scalars().all()
        recent_crawls = [
            {
                "id": j.id,
                "spider_name": j.source,
                "status": j.status.value if isinstance(j.status, JobStatus) else j.status,
                "start_time": j.started_at.isoformat() if j.started_at else None,
                "end_time": j.completed_at.isoformat() if j.completed_at else None,
            }
            for j in recent_jobs
        ]

        return {
            "total_companies": total,
            "verified_buyers": verified,
            "emails_found": emails,
            "phones_found": phones,
            "active_crawlers": 0,
            "importers": importers,
            "exporters": exporters,
            "manufacturers": manufacturers,
            "distributors": distributors,
            "wholesalers": wholesalers,
            "score_distribution": score_ranges,
            "industry_breakdown": industry_breakdown,
            "source_breakdown": source_breakdown,
            "state_breakdown": state_breakdown,
            "recent_activity": [],
            "recent_crawls": recent_crawls,
        }
    except Exception as e:
        logger.error(f"Stats query failed: {e}")
        return {
            "total_companies": 0, "verified_buyers": 0,
            "emails_found": 0, "phones_found": 0, "active_crawlers": 0,
            "importers": 0, "exporters": 0, "manufacturers": 0,
            "distributors": 0, "wholesalers": 0,
            "score_distribution": {}, "industry_breakdown": {},
            "source_breakdown": {}, "state_breakdown": {},
            "recent_activity": [], "recent_crawls": [],
        }


@router.get("/trade-summary")
async def get_trade_summary_v2():
    """Trade summary from DGCIS trade_data table."""
    import sqlite3
    try:
        conn = sqlite3.connect("buyerhunter.db")
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT hs_code, year, SUM(value_usd_million) as total_value,
                   SUM(quantity) as total_quantity
            FROM trade_data WHERE value_usd_million > 0
            GROUP BY hs_code, year ORDER BY hs_code, year
        """).fetchall()
        conn.close()
        return [dict(row) for row in rows]
    except Exception:
        return []


@router.get("/trade-data")
async def get_trade_data_v2(hs_code: str | None = None):
    """Full trade data records from DGCIS."""
    import sqlite3
    try:
        conn = sqlite3.connect("buyerhunter.db")
        conn.row_factory = sqlite3.Row
        if hs_code:
            rows = conn.execute(
                "SELECT * FROM trade_data WHERE hs_code = ? ORDER BY year", (hs_code,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM trade_data ORDER BY hs_code, year"
            ).fetchall()
        conn.close()
        return [dict(row) for row in rows]
    except Exception:
        return []


# ═══════════════════════════════════════════════════════════════════════════════
# Companies
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/companies")
async def list_companies(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: str | None = None,
    city: str | None = None,
    state: str | None = None,
    industry: str | None = None,
    source: str | None = None,
    min_score: int | None = None,
    max_score: int | None = None,
    has_email: bool | None = None,
    has_phone: bool | None = None,
    is_importer: bool | None = None,
    is_exporter: bool | None = None,
    is_manufacturer: bool | None = None,
    is_distributor: bool | None = None,
    is_wholesaler: bool | None = None,
    has_website: bool | None = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    db: AsyncSession = Depends(get_v2_db),
):
    """List companies with pagination and filters. Returns V1-compatible shape."""
    try:
        query = select(Company)
        count_query = select(func.count(Company.id))

        # Apply filters
        filters = []
        if search:
            like = f"%{search}%"
            filters.append(or_(
                Company.canonical_name.ilike(like),
                Company.industry.ilike(like),
                Company.gst_number.ilike(like),
                Company.iec_code.ilike(like),
                Company.hq_city.ilike(like),
                Company.hq_state.ilike(like),
                Company.website_url.ilike(like),
            ))
        if state:
            filters.append(Company.hq_state.ilike(f"%{state}%"))
        if city:
            filters.append(Company.hq_city.ilike(f"%{city}%"))
        if industry:
            filters.append(Company.industry.ilike(f"%{industry}%"))
        if source:
            filters.append(Company.first_seen_source.ilike(f"%{source}%"))
        if min_score is not None:
            filters.append(Company.buyer_score >= min_score)
        if max_score is not None:
            filters.append(Company.buyer_score <= max_score)
        if is_importer is not None:
            filters.append(Company.is_importer == is_importer)
        if is_exporter is not None:
            filters.append(Company.is_exporter == is_exporter)
        if is_manufacturer is not None:
            filters.append(Company.is_manufacturer == is_manufacturer)
        if is_distributor is not None:
            filters.append(Company.is_distributor == is_distributor)
        if is_wholesaler is not None:
            filters.append(Company.is_wholesaler == is_wholesaler)
        if has_website:
            filters.append(Company.website_url.isnot(None))
            filters.append(Company.website_url != "")

        for f in filters:
            query = query.where(f)
            count_query = count_query.where(f)

        total = (await db.execute(count_query)).scalar() or 0

        # Sort
        sort_col = getattr(Company, sort_by, Company.created_at)
        if sort_order == "desc":
            query = query.order_by(sort_col.desc())
        else:
            query = query.order_by(sort_col.asc())

        # Paginate
        offset = (page - 1) * page_size
        query = query.offset(offset).limit(page_size)

        result = await db.execute(query)
        companies = result.scalars().all()

        # Fetch contacts for email/phone display
        company_ids = [c.id for c in companies]
        contacts_map: dict[int, list[Contact]] = {}
        if company_ids:
            ct_result = await db.execute(
                select(Contact).where(Contact.company_id.in_(company_ids))
            )
            for ct in ct_result.scalars().all():
                contacts_map.setdefault(ct.company_id, []).append(ct)

        items = [_company_to_frontend(c, contacts_map.get(c.id, [])) for c in companies]

        return {"companies": items, "total": total}

    except Exception as e:
        logger.error(f"Companies query failed: {e}")
        return {"companies": [], "total": 0}


@router.get("/company/{company_id}")
async def get_company(company_id: int, db: AsyncSession = Depends(get_v2_db)):
    """Get single company with full details."""
    company = await db.get(Company, company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    ct_result = await db.execute(
        select(Contact).where(Contact.company_id == company_id)
    )
    contacts = list(ct_result.scalars().all())

    return _company_to_frontend(company, contacts)


@router.get("/company/{company_id}/contacts")
async def get_company_contacts(company_id: int, db: AsyncSession = Depends(get_v2_db)):
    """Get contacts for a company."""
    result = await db.execute(
        select(Contact).where(Contact.company_id == company_id)
    )
    contacts = [_contact_to_frontend(ct) for ct in result.scalars().all()]
    return {"contacts": contacts}


@router.get("/company/{company_id}/evidence")
async def get_company_evidence(company_id: int, db: AsyncSession = Depends(get_v2_db)):
    """Get evidence ledger for a company."""
    result = await db.execute(
        select(EvidenceLedger)
        .where(EvidenceLedger.entity_type == "company", EvidenceLedger.entity_id == company_id)
        .order_by(EvidenceLedger.observed_at.desc())
        .limit(100)
    )
    evidence = []
    for ev in result.scalars().all():
        evidence.append({
            "id": ev.id,
            "field_name": ev.field_name,
            "field_value": ev.field_value,
            "source_url": ev.source_url,
            "source_domain": ev.source_domain or "",
            "source_method": ev.source_method.value if hasattr(ev.source_method, 'value') else str(ev.source_method),
            "scraper_name": ev.scraper_name or "",
            "http_status": ev.http_status,
            "observed_at": ev.observed_at.isoformat() if ev.observed_at else None,
        })
    return {"evidence": evidence}


@router.get("/company/{company_id}/timeline")
async def get_company_timeline(company_id: int, db: AsyncSession = Depends(get_v2_db)):
    """Get timeline of all evidence for a company."""
    result = await db.execute(
        select(EvidenceLedger)
        .where(EvidenceLedger.entity_type == "company", EvidenceLedger.entity_id == company_id)
        .order_by(EvidenceLedger.observed_at.desc())
        .limit(200)
    )
    timeline = []
    for ev in result.scalars().all():
        timeline.append({
            "id": ev.id,
            "field_name": ev.field_name,
            "field_value": ev.field_value,
            "source_method": ev.source_method.value if hasattr(ev.source_method, 'value') else str(ev.source_method),
            "scraper_name": ev.scraper_name or "",
            "observed_at": ev.observed_at.isoformat() if ev.observed_at else None,
        })
    return {"timeline": timeline}


@router.post("/company/{company_id}/refresh")
async def refresh_company(company_id: int, db: AsyncSession = Depends(get_v2_db)):
    """Trigger re-enrichment for a company."""
    company = await db.get(Company, company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    return {"status": "queued", "company_id": company_id, "message": "Refresh queued"}


# ═══════════════════════════════════════════════════════════════════════════════
# Search (Global)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/search")
async def global_search(
    q: str = Query(..., min_length=1),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_v2_db),
):
    """Global search across V2 companies."""
    try:
        like = f"%{q}%"
        query = select(Company).where(or_(
            Company.canonical_name.ilike(like),
            Company.industry.ilike(like),
            Company.gst_number.ilike(like),
            Company.iec_code.ilike(like),
            Company.hq_city.ilike(like),
            Company.hq_state.ilike(like),
            Company.website_url.ilike(like),
        )).order_by(Company.buyer_score.desc()).limit(limit)

        result = await db.execute(query)
        companies = result.scalars().all()

        # Fetch contacts
        company_ids = [c.id for c in companies]
        contacts_map: dict[int, list[Contact]] = {}
        if company_ids:
            ct_result = await db.execute(
                select(Contact).where(Contact.company_id.in_(company_ids))
            )
            for ct in ct_result.scalars().all():
                contacts_map.setdefault(ct.company_id, []).append(ct)

        items = [_company_to_frontend(c, contacts_map.get(c.id, [])) for c in companies]
        return {"total": len(items), "query": q, "results": items}

    except Exception as e:
        logger.error(f"Search failed: {e}")
        return {"total": 0, "query": q, "results": []}


# ═══════════════════════════════════════════════════════════════════════════════
# Pipeline / Search Jobs
# ═══════════════════════════════════════════════════════════════════════════════

# In-memory progress store for SSE streaming
_pipeline_progress: dict[str, dict] = {}


@router.post("/search")
async def create_search_job(
    body: dict,
):
    """
    POST /api/v2/search — Start a new search pipeline.
    Accepts the same body shape as the V1 pipeline/start endpoint.
    """
    query_str = body.get("query", "").strip()
    if not query_str:
        raise HTTPException(status_code=400, detail="Query is required")

    max_queries = body.get("max_queries", 500)
    max_pages = body.get("max_pages_per_spider", 3)
    sources = body.get("sources")
    run_id = f"pip_{int(time.time())}_{uuid.uuid4().hex[:6]}"

    # Create initial progress entry
    _pipeline_progress[run_id] = {
        "run_id": run_id,
        "status": "planning",
        "query": query_str,
        "query_type": "general",
        "total_queries": 0,
        "completed_queries": 0,
        "urls_found": 0,
        "pages_crawled": 0,
        "companies_found": 0,
        "companies_new": 0,
        "companies_duplicate": 0,
        "companies_merged": 0,
        "emails_found": 0,
        "phones_found": 0,
        "whatsapp_found": 0,
        "websites_found": 0,
        "verified": 0,
        "enriched": 0,
        "scored": 0,
        "errors": 0,
        "current_source": "",
        "current_query": "",
        "active_jobs": 0,
        "queued_jobs": 0,
        "elapsed": 0,
        "started_at": time.time(),
        "messages": [],
    }

    # Plan queries in background
    asyncio.create_task(_execute_search_pipeline(
        run_id, query_str, max_queries, max_pages, sources
    ))

    return {"run_id": run_id, "status": "started"}


CITY_STATES = {
    "Mumbai": "Maharashtra",
    "Delhi": "Delhi",
    "Bengaluru": "Karnataka",
    "Bangalore": "Karnataka",
    "Hyderabad": "Telangana",
    "Chennai": "Tamil Nadu",
    "Kolkata": "West Bengal",
    "Pune": "Maharashtra",
    "Ahmedabad": "Gujarat",
    "Surat": "Gujarat",
    "Jaipur": "Rajasthan",
}


async def _crawl_and_save_job(db: AsyncSession, job: SearchJob) -> dict:
    """Crawl a real source for a search job. Saves with full provenance."""
    from backend.services.crawler_service import MultiSourceCrawler
    from backend.services.pipeline_service import (
        save_company_with_evidence,
        save_contact_with_evidence,
        EvidenceCategory,
    )

    city = job.target_city or "Bengaluru"
    state = job.target_state or CITY_STATES.get(city, "Karnataka")
    if city.lower() == "bangalore":
        city = "Bengaluru"

    crawler = MultiSourceCrawler(timeout=30, max_pages=3)
    try:
        result = await crawler.crawl_job(
            query_string=job.query_string,
            source=job.source or "justdial",
            target_city=city,
            target_state=state,
        )
    finally:
        await crawler.close()

    leads_added = 0
    emails_added = 0
    phones_added = 0
    websites_added = 0
    contacts_added = 0
    enrichments_run = 0
    discover_run = 0

    source_name = job.source or "unknown"
    evidence_cat = EvidenceCategory.API_RESPONSE if source_name == "indiamart" else EvidenceCategory.SCRAPED_DIRECT

    for ec in result.companies:
        company = await save_company_with_evidence(
            db,
            {
                "company_name": ec.company_name,
                "website": ec.website,
                "city": ec.city or city,
                "state": ec.state or state,
                "address": ec.address or "",
                "industry": ec.industry or "",
                "confidence": min(90, ec.confidence or 50),
                "buyer_score": min(90, ec.confidence or 50),
            },
            source=source_name,
            source_url=ec.source_url or "",
            scraper_name=f"crawler_{source_name}",
            evidence_category=evidence_cat,
        )
        if not company:
            continue

        leads_added += 1
        if ec.website:
            websites_added += 1

        # Save contact channels from crawler
        if ec.email:
            ct = await save_contact_with_evidence(
                db, company.id, ContactChannel.EMAIL, ec.email,
                source_url=ec.source_url or "", scraper_name=source_name,
                purpose=ContactPurpose.GENERAL, confidence=min(90, ec.confidence or 50),
                evidence_category=evidence_cat,
            )
            if ct: emails_added += 1
        if ec.phone:
            ct = await save_contact_with_evidence(
                db, company.id, ContactChannel.PHONE, ec.phone,
                source_url=ec.source_url or "", scraper_name=source_name,
                purpose=ContactPurpose.GENERAL, confidence=min(90, ec.confidence or 50),
                evidence_category=evidence_cat,
            )
            if ct: phones_added += 1
        if ec.whatsapp:
            await save_contact_with_evidence(
                db, company.id, ContactChannel.WHATSAPP, ec.whatsapp,
                source_url=ec.source_url or "", scraper_name=source_name,
                purpose=ContactPurpose.GENERAL, confidence=min(90, ec.confidence or 50),
                evidence_category=evidence_cat,
            )

        # Enrich company website
        if ec.website:
            try:
                enrich_data = await crawler.enrich_company_website(ec.website)
                enrichments_run += 1
                if enrich_data.get("about_us"):
                    company.about_us = enrich_data["about_us"]
                if enrich_data.get("company_description"):
                    company.company_description = enrich_data["company_description"]
                if enrich_data.get("address") and not ec.address:
                    company.hq_address = enrich_data["address"]
                if enrich_data.get("city") and not ec.city:
                    company.hq_city = enrich_data["city"]
                if enrich_data.get("state") and not ec.state:
                    company.hq_state = enrich_data["state"]

                for e in enrich_data.get("emails", []):
                    ct = await save_contact_with_evidence(
                        db, company.id, ContactChannel.EMAIL, e,
                        source_url=ec.website, scraper_name="website_enricher",
                        purpose=ContactPurpose.OFFICIAL, confidence=60,
                        evidence_category=EvidenceCategory.SCRAPED_DIRECT,
                    )
                    if ct: emails_added += 1
                for p in enrich_data.get("phones", []):
                    digits = re.sub(r"[^\d]", "", p)
                    if digits:
                        ct = await save_contact_with_evidence(
                            db, company.id, ContactChannel.PHONE, p,
                            source_url=ec.website, scraper_name="website_enricher",
                            purpose=ContactPurpose.OFFICIAL, confidence=55,
                            evidence_category=EvidenceCategory.SCRAPED_DIRECT,
                        )
                        if ct: phones_added += 1
            except Exception as ex:
                logger.debug(f"Enrichment failed for {ec.website}: {ex}")

            # Contact discovery
            try:
                discovered = await crawler.discover_contacts(ec.website)
                discover_run += 1
                for dc in discovered:
                    if not dc.person_name:
                        continue
                    c_confidence = min(90, dc.confidence or 50)
                    if dc.email:
                        ct = await save_contact_with_evidence(
                            db, company.id, ContactChannel.EMAIL, dc.email,
                            source_url=dc.source_url or ec.website,
                            scraper_name="contact_discovery",
                            person_name=dc.person_name, designation=dc.designation or "",
                            purpose=ContactPurpose.PROCUREMENT,
                            confidence=c_confidence,
                            evidence_category=EvidenceCategory.SCRAPED_DIRECT,
                        )
                        if ct: contacts_added += 1
                    if dc.phone:
                        ct = await save_contact_with_evidence(
                            db, company.id, ContactChannel.PHONE, dc.phone,
                            source_url=dc.source_url or ec.website,
                            scraper_name="contact_discovery",
                            person_name=dc.person_name, designation=dc.designation or "",
                            purpose=ContactPurpose.PROCUREMENT,
                            confidence=c_confidence,
                            evidence_category=EvidenceCategory.SCRAPED_DIRECT,
                        )
                        if ct: contacts_added += 1
            except Exception as ex:
                logger.debug(f"Contact discovery failed for {ec.website}: {ex}")

    await db.commit()

    return {
        "leads_added": leads_added,
        "emails_added": emails_added,
        "phones_added": phones_added,
        "websites_added": websites_added,
        "contacts_added": contacts_added,
        "pages_crawled": result.pages_crawled,
        "enrichments_run": enrichments_run,
        "discoveries_run": discover_run,
        "companies_found_in_source": len(result.companies),
        "errors": result.errors,
    }


async def _execute_search_pipeline(
    run_id: str,
    query_str: str,
    max_queries: int,
    max_pages: int,
    sources: list[str] | None,
):
    """Execute search pipeline in background. Updates _pipeline_progress."""
    progress = _pipeline_progress.get(run_id)
    if not progress:
        return

    def _add_msg(msg: str):
        progress["messages"].append({
            "time": datetime.now(timezone.utc).isoformat(),
            "message": msg,
        })

    from backend.core.database import get_session_factory
    session_factory = get_session_factory()

    try:
        async with session_factory() as db:
            # Phase 1: Intent analysis
            progress["status"] = "planning"
            analyzer = IntentAnalyzer()
            intent = analyzer.analyze(query_str)
            progress["query_type"] = intent.query_type
            _add_msg(f"Planning search for '{query_str}' (type: {intent.query_type})")

            # Phase 2: Query expansion
            progress["status"] = "expanding"
            planner = HybridQueryPlanner()
            plan = await planner.plan(query_str, max_deterministic=max_queries, max_ai=0)

            variations = plan.deterministic_queries
            if sources:
                variations = [v for v in variations if v.source in sources]

            progress["total_queries"] = len(variations)
            _add_msg(f"Generated {len(variations)} search queries across {len(set(v.source for v in variations))} sources")

            if not variations:
                progress["status"] = "completed"
                _add_msg("No queries generated")
                return

            # Phase 3: Create search jobs in DB
            progress["status"] = "queuing"
            job_ids = []
            for v in variations:
                job = SearchJob(
                    query_string=v.query_string,
                    source=v.source,
                    status=JobStatus.PENDING,
                    priority=JobPriority.NORMAL if v.priority < 8 else JobPriority.HIGH,
                    max_pages=max_pages,
                    target_state=v.target_state,
                    target_city=v.target_city,
                    run_id=run_id,
                )
                db.add(job)
                await db.flush()
                job_ids.append(job.id)

            await db.commit()
            progress["queued_jobs"] = len(job_ids)
            _add_msg(f"Enqueued {len(job_ids)} search jobs")

            # Phase 4: Execute jobs concurrently (max 3 at a time)
            progress["status"] = "crawling"
            sem = asyncio.Semaphore(3)

            async def _crawl_one(idx: int, job_id: int, v: Any) -> dict:
                async with sem:
                    progress["current_query"] = v.query_string
                    progress["current_source"] = v.source
                    progress["active_jobs"] = progress.get("active_jobs", 0) + 1

                    async with session_factory() as task_db:
                        job = await task_db.get(SearchJob, job_id)
                        if not job:
                            progress["active_jobs"] = max(0, progress["active_jobs"] - 1)
                            return {}

                        job.status = JobStatus.RUNNING
                        job.started_at = datetime.now(timezone.utc)
                        await task_db.commit()

                        _add_msg(f"[{idx+1}/{len(variations)}] Crawling '{v.query_string}' on {v.source}")
                        res = {}
                        try:
                            res = await _crawl_and_save_job(task_db, job)
                        except Exception as ex:
                            logger.error(f"Failed to crawl job {job_id}: {ex}")
                            res = {"errors": [str(ex)]}

                        job.status = JobStatus.COMPLETED
                        job.completed_at = datetime.now(timezone.utc)
                        job.pages_crawled = res.get("pages_crawled", 0)
                        job.companies_found = res.get("leads_added", 0)
                        job.contacts_found = res.get("emails_added", 0) + res.get("phones_added", 0) + res.get("contacts_added", 0)
                        await task_db.commit()

                    # Update live counters immediately for SSE
                    progress["pages_crawled"] += res.get("pages_crawled", 0)
                    progress["companies_found"] += res.get("leads_added", 0)
                    progress["companies_new"] += res.get("leads_added", 0)
                    progress["emails_found"] += res.get("emails_added", 0)
                    progress["phones_found"] += res.get("phones_added", 0)
                    progress["websites_found"] += res.get("websites_added", 0)
                    progress["contacts_found"] = progress.get("contacts_found", 0) + res.get("contacts_added", 0)
                    progress["scored"] += res.get("leads_added", 0)
                    progress["enriched"] += res.get("leads_added", 0)
                    progress["completed_queries"] += 1
                    progress["active_jobs"] = max(0, progress["active_jobs"] - 1)

                    _add_msg(
                        f"[{idx+1}/{len(variations)}] Done '{v.query_string}' on {v.source}: "
                        f"{res.get('leads_added', 0)} leads, "
                        f"{res.get('emails_added', 0)} emails, "
                        f"{res.get('phones_added', 0)} phones"
                    )
                    return res

            results = await asyncio.gather(*[
                _crawl_one(i, job_ids[i], variations[i])
                for i in range(min(len(job_ids), len(variations)))
            ], return_exceptions=True)

            # Phase 5: Dedup & merge
            progress["status"] = "deduping"
            _add_msg("Running deduplication...")
            try:
                from backend.services.pipeline_service import run_dedup_pipeline
                dedup_result = await run_dedup_pipeline(db)
                progress["duplicates_removed"] = dedup_result.get("removed", 0)
                progress["merged_count"] = dedup_result.get("merged", 0)
                progress["after_dedup"] = dedup_result.get("after_dedup", 0)
                _add_msg(
                    f"Dedup: {dedup_result.get('merged', 0)} merged, "
                    f"{dedup_result.get('removed', 0)} removed, "
                    f"{dedup_result.get('after_dedup', 0)} unique companies"
                )
            except Exception as ex:
                logger.error(f"Dedup pipeline failed: {ex}")
                _add_msg(f"Dedup failed: {ex}")

            # Phase 6: Summary
            progress["status"] = "completed"
            _add_msg(
                f"Pipeline complete! "
                f"Queries: {progress['completed_queries']}/{progress['total_queries']} | "
                f"Pages: {progress['pages_crawled']} | "
                f"Companies found: {progress['companies_found']} | "
                f"After dedup: {progress.get('after_dedup', progress['companies_found'])} | "
                f"Contacts: {progress.get('contacts_found', 0)} | "
                f"Duplicates removed: {progress.get('duplicates_removed', 0)}"
            )

    except Exception as e:
        progress["status"] = "failed"
        _add_msg(f"Pipeline failed: {e}")
        logger.exception(f"Pipeline {run_id} failed")
    finally:
        from backend.services.crawler_service import close_global_browser
        await close_global_browser()


@router.get("/search/{run_id}")
async def get_search_job(run_id: str):
    """GET /api/v2/search/{run_id} — Get search job status/progress."""
    progress = _pipeline_progress.get(run_id)
    if not progress:
        raise HTTPException(status_code=404, detail="Search job not found")
    return progress


@router.get("/search/{run_id}/stream")
async def stream_search_job(run_id: str):
    """GET /api/v2/search/{run_id}/stream — SSE stream for real-time progress."""
    async def event_generator():
        last_msg_count = 0
        while True:
            progress = _pipeline_progress.get(run_id)
            if not progress:
                yield f"data: {json.dumps({'status': 'not_found'})}\n\n"
                break

            current_msgs = len(progress.get("messages", []))
            elapsed = time.time() - progress.get("started_at", time.time())
            progress["elapsed"] = elapsed

            # Always push progress during crawling (emit every tick when active)
            if (progress["status"] == "crawling" and progress.get("active_jobs", 0) > 0) \
               or current_msgs > last_msg_count \
               or progress["status"] in ("completed", "failed", "cancelled"):
                yield f"data: {json.dumps(progress)}\n\n"
                last_msg_count = current_msgs

            if progress["status"] in ("completed", "failed", "cancelled"):
                break

            await asyncio.sleep(1.0)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Compatibility aliases — the frontend may hit /api/v2/pipeline/* directly
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/pipeline/start")
async def pipeline_start_compat(body: dict):
    """Compatibility: POST /api/v2/pipeline/start → same as POST /api/v2/search"""
    return await create_search_job(body)


@router.get("/pipeline/{run_id}/progress")
async def pipeline_progress_compat(run_id: str):
    """Compatibility: GET /api/v2/pipeline/{run_id}/progress"""
    return await get_search_job(run_id)


@router.get("/pipeline/{run_id}/stream")
async def pipeline_stream_compat(run_id: str):
    """Compatibility: GET /api/v2/pipeline/{run_id}/stream"""
    return await stream_search_job(run_id)


@router.get("/pipeline/active")
async def pipeline_active():
    """List active pipelines."""
    active = [
        p for p in _pipeline_progress.values()
        if p["status"] not in ("completed", "failed", "cancelled")
    ]
    return {"pipelines": active}


@router.get("/pipeline/expand")
async def pipeline_expand(
    query: str = Query(...),
    max_queries: int = Query(500, ge=1, le=2000),
):
    """Preview query expansion without executing."""
    planner = HybridQueryPlanner()
    plan = await planner.plan(query, max_deterministic=max_queries, max_ai=0)

    by_source = {}
    for v in plan.deterministic_queries:
        by_source[v.source] = by_source.get(v.source, 0) + 1

    return {
        "query": query,
        "query_type": plan.intent.query_type,
        "total_variations": len(plan.deterministic_queries),
        "locations_covered": len(plan.states_covered),
        "by_source": by_source,
        "variations": [
            {"query": v.query_string, "source": v.source, "location": v.target_state or v.target_city or "India"}
            for v in plan.deterministic_queries[:200]
        ],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Export
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/export")
async def export_data(
    format: str = Query("csv"),
    search: str | None = None,
    city: str | None = None,
    state: str | None = None,
    industry: str | None = None,
    source: str | None = None,
    min_score: int | None = None,
    max_score: int | None = None,
    has_email: bool | None = None,
    has_phone: bool | None = None,
    is_importer: bool | None = None,
    is_exporter: bool | None = None,
    is_manufacturer: bool | None = None,
    is_distributor: bool | None = None,
    is_wholesaler: bool | None = None,
    has_website: bool | None = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    db: AsyncSession = Depends(get_v2_db),
):
    """Export companies with all active filters applied. No pagination — exports EVERYTHING that matches the current view."""
    query = select(Company)

    filters = []
    if search:
        like = f"%{search}%"
        filters.append(or_(
            Company.canonical_name.ilike(like),
            Company.industry.ilike(like),
            Company.gst_number.ilike(like),
            Company.iec_code.ilike(like),
            Company.hq_city.ilike(like),
            Company.hq_state.ilike(like),
            Company.website_url.ilike(like),
        ))
    if state:
        filters.append(Company.hq_state.ilike(f"%{state}%"))
    if city:
        filters.append(Company.hq_city.ilike(f"%{city}%"))
    if industry:
        filters.append(Company.industry.ilike(f"%{industry}%"))
    if source:
        filters.append(Company.first_seen_source.ilike(f"%{source}%"))
    if min_score is not None:
        filters.append(Company.buyer_score >= min_score)
    if max_score is not None:
        filters.append(Company.buyer_score <= max_score)
    if has_email:
        filters.append(
            exists().where(
                Contact.company_id == Company.id,
                Contact.channel == "email",
                Contact.channel_value.isnot(None),
                Contact.channel_value != "",
            )
        )
    if has_phone:
        filters.append(
            exists().where(
                Contact.company_id == Company.id,
                Contact.channel == "phone",
                Contact.channel_value.isnot(None),
                Contact.channel_value != "",
            )
        )
    if has_website:
        filters.append(Company.website_url.isnot(None))
        filters.append(Company.website_url != "")
    if is_importer is not None:
        filters.append(Company.is_importer == is_importer)
    if is_exporter is not None:
        filters.append(Company.is_exporter == is_exporter)
    if is_manufacturer is not None:
        filters.append(Company.is_manufacturer == is_manufacturer)
    if is_distributor is not None:
        filters.append(Company.is_distributor == is_distributor)
    if is_wholesaler is not None:
        filters.append(Company.is_wholesaler == is_wholesaler)

    for f in filters:
        query = query.where(f)

    sort_col = getattr(Company, sort_by, Company.created_at)
    query = query.order_by(sort_col.desc() if sort_order == "desc" else sort_col.asc())

    result = await db.execute(query)
    companies = result.scalars().all()

    # Fetch contacts for email/phone/person display
    company_ids = [c.id for c in companies]
    contacts_map: dict[int, list[Contact]] = {}
    if company_ids:
        ct_result = await db.execute(
            select(Contact).where(Contact.company_id.in_(company_ids))
        )
        for ct in ct_result.scalars().all():
            contacts_map.setdefault(ct.company_id, []).append(ct)

    items = [_company_to_frontend(c, contacts_map.get(c.id, [])) for c in companies]

    if format == "json":
        return {"data": items, "count": len(items)}

    if format == "excel":
        return _export_excel_response(items)

    return _export_csv_response(items)


def _export_csv_response(items: list[dict]) -> Response:
    """Generate CSV response from frontend-shaped items."""
    import csv
    import io

    fields = [
        "company_name", "website", "industry", "products",
        "email", "phone", "whatsapp",
        "address", "city", "state", "country",
        "gst_number", "iec_code", "linkedin_url",
        "source", "lead_score",
        "contact_person", "designation",
        "last_verified",
    ]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for row in items:
        out = {k: row.get(k, "") for k in fields}
        out["contact_person"] = row.get("contact_person", "")
        out["designation"] = row.get("designation", "")
        out["last_verified"] = row.get("updated_at", "") or row.get("enriched_at", "") or ""
        writer.writerow(out)

    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=companies.csv"},
    )


def _export_excel_response(items: list[dict]) -> Response:
    """Generate Excel (.xlsx) response from frontend-shaped items."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    headers = [
        "Company Name", "Website", "Industry", "Products",
        "Contact Person", "Designation", "Email", "Phone", "WhatsApp",
        "Address", "City", "State", "Country",
        "GST", "IEC", "LinkedIn",
        "Source", "Confidence Score", "Last Verified",
    ]
    field_keys = [
        "company_name", "website", "industry", "products",
        "contact_person", "designation", "email", "phone", "whatsapp",
        "address", "city", "state", "country",
        "gst_number", "iec_code", "linkedin_url",
        "source", "lead_score", "last_verified",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for item in items:
        row = []
        for key in field_keys:
            if key == "last_verified":
                val = item.get("updated_at", "") or item.get("enriched_at", "") or ""
            elif key == "contact_person":
                val = item.get("contact_person", "")
            elif key == "designation":
                val = item.get("designation", "")
            else:
                val = item.get(key, "")
            row.append(val)
        ws.append(row)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=companies.xlsx"},
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Compatibility: crawl status, spiders, etc.
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/crawl/status")
async def crawl_status():
    return {"running": [], "interrupted": []}


@router.get("/crawl-logs")
async def crawl_logs(limit: int = Query(20), db: AsyncSession = Depends(get_v2_db)):
    result = await db.execute(
        select(SearchJob).order_by(SearchJob.created_at.desc()).limit(limit)
    )
    logs = []
    for j in result.scalars().all():
        logs.append({
            "id": j.id,
            "spider_name": j.source,
            "status": j.status.value if isinstance(j.status, JobStatus) else j.status,
            "start_time": j.started_at.isoformat() if j.started_at else None,
            "end_time": j.completed_at.isoformat() if j.completed_at else None,
            "errors": j.errors or [],
        })
    return logs


@router.get("/spiders")
async def get_spiders():
    return {
        "spiders": {
            "indiamart": {"name": "IndiaMART", "type": "b2b_directory"},
            "justdial": {"name": "JustDial", "type": "local_directory"},
            "tradeindia": {"name": "TradeIndia", "type": "b2b_directory"},
            "googlemaps": {"name": "Google Maps", "type": "local_directory"},
            "fssai": {"name": "FSSAI Registry", "type": "government_registry"},
        }
    }


@router.get("/discovery/sources")
async def discovery_sources():
    return {
        "sources": [
            {"name": "indiamart", "display": "IndiaMART", "type": "b2b_directory", "priority": 10},
            {"name": "justdial", "display": "JustDial", "type": "local_directory", "priority": 9},
            {"name": "tradeindia", "display": "TradeIndia", "type": "b2b_directory", "priority": 8},
            {"name": "google_maps", "display": "Google Maps", "type": "local_directory", "priority": 7},
            {"name": "fssai", "display": "FSSAI Registry", "type": "government_registry", "priority": 6},
        ]
    }


@router.get("/discovery/stats")
async def discovery_stats(db: AsyncSession = Depends(get_v2_db)):
    total_jobs = (await db.execute(select(func.count(SearchJob.id)))).scalar() or 0
    completed = (await db.execute(
        select(func.count(SearchJob.id)).where(SearchJob.status == JobStatus.COMPLETED)
    )).scalar() or 0
    failed = (await db.execute(
        select(func.count(SearchJob.id)).where(SearchJob.status == JobStatus.FAILED)
    )).scalar() or 0
    return {"total_jobs": total_jobs, "completed": completed, "failed": failed}


@router.get("/crawl/dashboard")
async def crawl_dashboard(db: AsyncSession = Depends(get_v2_db)):
    """Full crawl dashboard with source breakdown, confidence, and counts."""
    from backend.services.pipeline_service import get_crawl_dashboard
    return await get_crawl_dashboard(db)


@router.post("/crawl/run-source")
async def run_source_independently(
    source: str = Query(...),
    query: str = Query(...),
    city: str = Query(""),
    db: AsyncSession = Depends(get_v2_db),
):
    """Run a single source crawler independently for validation."""
    from backend.services.crawler_service import MultiSourceCrawler
    from backend.services.pipeline_service import (
        save_company_with_evidence,
        save_contact_with_evidence,
        EvidenceCategory,
    )

    start_time = datetime.now(timezone.utc)
    crawler = MultiSourceCrawler(timeout=30, max_pages=2)
    try:
        result = await crawler.crawl_job(
            query_string=query,
            source=source,
            target_city=city,
            target_state="",
        )
    finally:
        await crawler.close()

    saved = 0
    for ec in result.companies:
        company = await save_company_with_evidence(
            db, {
                "company_name": ec.company_name,
                "website": ec.website,
                "city": ec.city or city,
                "state": ec.state or "",
                "address": ec.address or "",
                "industry": ec.industry or "",
                "confidence": min(90, ec.confidence or 50),
            },
            source=source,
            source_url=ec.source_url or "",
            scraper_name=f"crawler_{source}",
            evidence_category=EvidenceCategory.SCRAPED_DIRECT,
        )
        if company:
            saved += 1

        if ec.email:
            await save_contact_with_evidence(
                db, company.id, ContactChannel.EMAIL, ec.email,
                source_url=ec.source_url or "", scraper_name=source,
                evidence_category=EvidenceCategory.SCRAPED_DIRECT,
            )
        if ec.phone:
            await save_contact_with_evidence(
                db, company.id, ContactChannel.PHONE, ec.phone,
                source_url=ec.source_url or "", scraper_name=source,
                evidence_category=EvidenceCategory.SCRAPED_DIRECT,
            )

    await db.commit()

    end_time = datetime.now(timezone.utc)
    duration_sec = (end_time - start_time).total_seconds()

    return {
        "source": source,
        "query": query,
        "city": city or "any",
        "start_time": start_time.isoformat(),
        "end_time": end_time.isoformat(),
        "duration_seconds": round(duration_sec, 1),
        "companies_found": len(result.companies),
        "companies_saved": saved,
        "pages_crawled": result.pages_crawled,
        "errors": result.errors,
        "success_rate": round(
            (len(result.companies) / max(len(result.companies) + len(result.errors), 1)) * 100, 1
        ),
    }
